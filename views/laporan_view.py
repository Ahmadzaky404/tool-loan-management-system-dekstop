import customtkinter as ctk
import tkinter.messagebox as msg
from datetime import date, timedelta
from config import *
from components.widgets import DataTable, SectionHeader, DatePicker
from db.laporan_db import get_laporan_peminjaman, get_laporan_pengembalian, get_laporan_denda
from utils.format import tgl, rupiah
import os


class LaporanView(ctk.CTkFrame):
    def __init__(self, parent, user: dict):
        super().__init__(parent, fg_color="transparent")
        self.user = user
        self._build()

    def _build(self):
        # Header
        header = ctk.CTkFrame(self, fg_color="transparent")
        header.pack(fill="x", pady=(0, 12))
        SectionHeader(header, "Laporan").pack(side="left", fill="x", expand=True)

        # Tab view
        self.tabview = ctk.CTkTabview(self, fg_color=CARD_BG, corner_radius=12)
        self.tabview.pack(fill="both", expand=True)

        # Tabs
        self.tab_peminjaman = self.tabview.add("📋 Peminjaman")
        self.tab_pengembalian = self.tabview.add("↩️ Pengembalian")
        self.tab_denda = self.tabview.add("💰 Denda")

        self._build_peminjaman_tab()
        self._build_pengembalian_tab()
        self._build_denda_tab()
        
        # Load data awal
        self.after(100, self._load_peminjaman)

    def _build_peminjaman_tab(self):
        # Header info
        info_frame = ctk.CTkFrame(self.tab_peminjaman, fg_color="transparent")
        info_frame.pack(fill="x", pady=(8, 12))
        ctk.CTkLabel(info_frame, text="Laporan data peminjaman alat", 
                     text_color=TEXT_MUTED, font=ctk.CTkFont(size=12)).pack(anchor="w")

        # Filter card
        filter_card = ctk.CTkFrame(self.tab_peminjaman, fg_color=CARD_BG, corner_radius=12)
        filter_card.pack(fill="x", pady=(0, 12), padx=2)
        
        filter_content = ctk.CTkFrame(filter_card, fg_color="transparent")
        filter_content.pack(fill="x", padx=16, pady=16)

        # Row 1: Tanggal
        row1 = ctk.CTkFrame(filter_content, fg_color="transparent")
        row1.pack(fill="x", pady=(0, 12))
        
        col1 = ctk.CTkFrame(row1, fg_color="transparent")
        col1.pack(side="left", fill="x", expand=True, padx=(0, 8))
        ctk.CTkLabel(col1, text="Dari Tanggal", text_color=TEXT_LABEL, 
                     font=ctk.CTkFont(size=11)).pack(anchor="w", pady=(0, 4))
        self.pem_tgl_dari = DatePicker(col1)
        self.pem_tgl_dari.pack(fill="x")
        self.pem_tgl_dari.set((date.today() - timedelta(days=30)).strftime('%Y-%m-%d'))

        col2 = ctk.CTkFrame(row1, fg_color="transparent")
        col2.pack(side="left", fill="x", expand=True, padx=(0, 8))
        ctk.CTkLabel(col2, text="Sampai Tanggal", text_color=TEXT_LABEL,
                     font=ctk.CTkFont(size=11)).pack(anchor="w", pady=(0, 4))
        self.pem_tgl_sampai = DatePicker(col2)
        self.pem_tgl_sampai.pack(fill="x")

        col3 = ctk.CTkFrame(row1, fg_color="transparent")
        col3.pack(side="left", fill="x", expand=True)
        ctk.CTkLabel(col3, text="Status", text_color=TEXT_LABEL,
                     font=ctk.CTkFont(size=11)).pack(anchor="w", pady=(0, 4))
        self.pem_status = ctk.CTkOptionMenu(col3, 
                                             values=["semua", "pending", "disetujui", "ditolak", "selesai"],
                                             fg_color=INPUT_BG, button_color=PRIMARY)
        self.pem_status.pack(fill="x")

        # Row 2: Buttons
        row2 = ctk.CTkFrame(filter_content, fg_color="transparent")
        row2.pack(fill="x")
        
        ctk.CTkButton(row2, text="🔍  FILTER", width=140, height=38,
                      fg_color=DANGER, hover_color="#991B1B",
                      font=ctk.CTkFont(size=12, weight="bold"),
                      command=self._load_peminjaman).pack(side="right")

        # Export buttons
        export_frame = ctk.CTkFrame(self.tab_peminjaman, fg_color="transparent")
        export_frame.pack(fill="x", pady=(0, 12))
        
        ctk.CTkButton(export_frame, text="📄  Export PDF", width=130, height=36,
                      fg_color=DANGER, hover_color="#991B1B",
                      font=ctk.CTkFont(size=12),
                      command=self._export_peminjaman_pdf).pack(side="left", padx=(0, 8))
        
        ctk.CTkButton(export_frame, text="📊  Export Excel", width=130, height=36,
                      fg_color=SUCCESS, hover_color="#166534",
                      font=ctk.CTkFont(size=12),
                      command=self._export_peminjaman_excel).pack(side="left")

        # Table
        cols = ["ID", "Peminjam", "Tgl Pinjam", "Tgl Kembali", "Status", "Alat"]
        widths = [50, 150, 100, 100, 100, 300]
        self.tbl_peminjaman = DataTable(self.tab_peminjaman, cols, widths)
        self.tbl_peminjaman.pack(fill="both", expand=True)

    def _build_pengembalian_tab(self):
        # Header info
        info_frame = ctk.CTkFrame(self.tab_pengembalian, fg_color="transparent")
        info_frame.pack(fill="x", pady=(8, 12))
        ctk.CTkLabel(info_frame, text="Laporan data pengembalian alat", 
                     text_color=TEXT_MUTED, font=ctk.CTkFont(size=12)).pack(anchor="w")

        # Filter card
        filter_card = ctk.CTkFrame(self.tab_pengembalian, fg_color=CARD_BG, corner_radius=12)
        filter_card.pack(fill="x", pady=(0, 12), padx=2)
        
        filter_content = ctk.CTkFrame(filter_card, fg_color="transparent")
        filter_content.pack(fill="x", padx=16, pady=16)

        # Row 1: Tanggal
        row1 = ctk.CTkFrame(filter_content, fg_color="transparent")
        row1.pack(fill="x", pady=(0, 12))
        
        col1 = ctk.CTkFrame(row1, fg_color="transparent")
        col1.pack(side="left", fill="x", expand=True, padx=(0, 8))
        ctk.CTkLabel(col1, text="Dari Tanggal", text_color=TEXT_LABEL,
                     font=ctk.CTkFont(size=11)).pack(anchor="w", pady=(0, 4))
        self.pen_tgl_dari = DatePicker(col1)
        self.pen_tgl_dari.pack(fill="x")
        self.pen_tgl_dari.set((date.today() - timedelta(days=30)).strftime('%Y-%m-%d'))

        col2 = ctk.CTkFrame(row1, fg_color="transparent")
        col2.pack(side="left", fill="x", expand=True, padx=(0, 8))
        ctk.CTkLabel(col2, text="Sampai Tanggal", text_color=TEXT_LABEL,
                     font=ctk.CTkFont(size=11)).pack(anchor="w", pady=(0, 4))
        self.pen_tgl_sampai = DatePicker(col2)
        self.pen_tgl_sampai.pack(fill="x")

        # Row 2: Button
        row2 = ctk.CTkFrame(filter_content, fg_color="transparent")
        row2.pack(fill="x")
        
        ctk.CTkButton(row2, text="🔍  FILTER", width=140, height=38,
                      fg_color=DANGER, hover_color="#991B1B",
                      font=ctk.CTkFont(size=12, weight="bold"),
                      command=self._load_pengembalian).pack(side="right")

        # Export buttons
        export_frame = ctk.CTkFrame(self.tab_pengembalian, fg_color="transparent")
        export_frame.pack(fill="x", pady=(0, 12))
        
        ctk.CTkButton(export_frame, text="📄  Export PDF", width=130, height=36,
                      fg_color=DANGER, hover_color="#991B1B",
                      font=ctk.CTkFont(size=12),
                      command=self._export_pengembalian_pdf).pack(side="left", padx=(0, 8))
        
        ctk.CTkButton(export_frame, text="📊  Export Excel", width=130, height=36,
                      fg_color=SUCCESS, hover_color="#166534",
                      font=ctk.CTkFont(size=12),
                      command=self._export_pengembalian_excel).pack(side="left")

        # Table
        cols = ["ID", "Peminjam", "Tgl Kembali", "Kondisi", "Denda", "Verifikator"]
        widths = [50, 150, 100, 100, 100, 150]
        self.tbl_pengembalian = DataTable(self.tab_pengembalian, cols, widths)
        self.tbl_pengembalian.pack(fill="both", expand=True)

    def _build_denda_tab(self):
        # Header info
        info_frame = ctk.CTkFrame(self.tab_denda, fg_color="transparent")
        info_frame.pack(fill="x", pady=(8, 12))
        ctk.CTkLabel(info_frame, text="Laporan data denda keterlambatan", 
                     text_color=TEXT_MUTED, font=ctk.CTkFont(size=12)).pack(anchor="w")

        # Filter card
        filter_card = ctk.CTkFrame(self.tab_denda, fg_color=CARD_BG, corner_radius=12)
        filter_card.pack(fill="x", pady=(0, 12), padx=2)
        
        filter_content = ctk.CTkFrame(filter_card, fg_color="transparent")
        filter_content.pack(fill="x", padx=16, pady=16)

        # Row 1: Tanggal dan Status
        row1 = ctk.CTkFrame(filter_content, fg_color="transparent")
        row1.pack(fill="x", pady=(0, 12))
        
        col1 = ctk.CTkFrame(row1, fg_color="transparent")
        col1.pack(side="left", fill="x", expand=True, padx=(0, 8))
        ctk.CTkLabel(col1, text="Dari Tanggal", text_color=TEXT_LABEL,
                     font=ctk.CTkFont(size=11)).pack(anchor="w", pady=(0, 4))
        self.den_tgl_dari = DatePicker(col1)
        self.den_tgl_dari.pack(fill="x")
        self.den_tgl_dari.set((date.today() - timedelta(days=30)).strftime('%Y-%m-%d'))

        col2 = ctk.CTkFrame(row1, fg_color="transparent")
        col2.pack(side="left", fill="x", expand=True, padx=(0, 8))
        ctk.CTkLabel(col2, text="Sampai Tanggal", text_color=TEXT_LABEL,
                     font=ctk.CTkFont(size=11)).pack(anchor="w", pady=(0, 4))
        self.den_tgl_sampai = DatePicker(col2)
        self.den_tgl_sampai.pack(fill="x")

        col3 = ctk.CTkFrame(row1, fg_color="transparent")
        col3.pack(side="left", fill="x", expand=True)
        ctk.CTkLabel(col3, text="Status", text_color=TEXT_LABEL,
                     font=ctk.CTkFont(size=11)).pack(anchor="w", pady=(0, 4))
        self.den_status = ctk.CTkOptionMenu(col3, 
                                             values=["semua", "belum_bayar", "lunas"],
                                             fg_color=INPUT_BG, button_color=PRIMARY)
        self.den_status.pack(fill="x")

        # Row 2: Button
        row2 = ctk.CTkFrame(filter_content, fg_color="transparent")
        row2.pack(fill="x")
        
        ctk.CTkButton(row2, text="🔍  FILTER", width=140, height=38,
                      fg_color=DANGER, hover_color="#991B1B",
                      font=ctk.CTkFont(size=12, weight="bold"),
                      command=self._load_denda).pack(side="right")

        # Export buttons
        export_frame = ctk.CTkFrame(self.tab_denda, fg_color="transparent")
        export_frame.pack(fill="x", pady=(0, 12))
        
        ctk.CTkButton(export_frame, text="📄  Export PDF", width=130, height=36,
                      fg_color=DANGER, hover_color="#991B1B",
                      font=ctk.CTkFont(size=12),
                      command=self._export_denda_pdf).pack(side="left", padx=(0, 8))
        
        ctk.CTkButton(export_frame, text="📊  Export Excel", width=130, height=36,
                      fg_color=SUCCESS, hover_color="#166534",
                      font=ctk.CTkFont(size=12),
                      command=self._export_denda_excel).pack(side="left")

        # Table
        cols = ["ID", "Peminjam", "Jumlah", "Status", "Tgl Bayar", "Keterangan"]
        widths = [50, 150, 100, 100, 100, 250]
        self.tbl_denda = DataTable(self.tab_denda, cols, widths)
        self.tbl_denda.pack(fill="both", expand=True)

    def _load_peminjaman(self):
        tgl_dari = self.pem_tgl_dari.get()
        tgl_sampai = self.pem_tgl_sampai.get()
        status = self.pem_status.get()
        
        self._data_peminjaman = get_laporan_peminjaman(tgl_dari, tgl_sampai, status)
        rows = [(
            d['id'],
            d['nama_peminjam'],
            tgl(d['tanggal_pinjam']),
            tgl(d['tanggal_kembali']),
            d['status'],
            d['alat_list'] or '-'
        ) for d in self._data_peminjaman]
        self.tbl_peminjaman.load(rows)

    def _load_pengembalian(self):
        tgl_dari = self.pen_tgl_dari.get()
        tgl_sampai = self.pen_tgl_sampai.get()
        
        self._data_pengembalian = get_laporan_pengembalian(tgl_dari, tgl_sampai)
        rows = [(
            d['id'],
            d['nama_peminjam'],
            tgl(d['tanggal_kembali_aktual']),
            d['kondisi_alat'],
            rupiah(d['total_denda']),
            d['nama_verifikator'] or 'Belum'
        ) for d in self._data_pengembalian]
        self.tbl_pengembalian.load(rows)

    def _load_denda(self):
        tgl_dari = self.den_tgl_dari.get()
        tgl_sampai = self.den_tgl_sampai.get()
        status = self.den_status.get()
        
        self._data_denda = get_laporan_denda(tgl_dari, tgl_sampai, status)
        rows = [(
            d['id'],
            d['nama_peminjam'],
            rupiah(d['jumlah']),
            d['status'],
            tgl(d.get('tanggal_bayar')),
            d.get('keterangan', '-')[:50]
        ) for d in self._data_denda]
        self.tbl_denda.load(rows)

    # Export Excel methods
    def _export_peminjaman_excel(self):
        if not hasattr(self, '_data_peminjaman'):
            msg.showwarning("Peringatan", "Tampilkan data terlebih dahulu")
            return
        
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Laporan Peminjaman"
            
            # Header
            headers = ["ID", "Peminjam", "Email", "Tgl Pinjam", "Tgl Kembali", "Status", "Approver", "Alat"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(1, col, header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Data
            for row, d in enumerate(self._data_peminjaman, 2):
                ws.cell(row, 1, d['id'])
                ws.cell(row, 2, d['nama_peminjam'])
                ws.cell(row, 3, d.get('email', '-'))
                ws.cell(row, 4, str(d['tanggal_pinjam']))
                ws.cell(row, 5, str(d['tanggal_kembali']))
                ws.cell(row, 6, d['status'])
                ws.cell(row, 7, d.get('nama_approver', '-'))
                ws.cell(row, 8, d.get('alat_list', '-'))
            
            # Auto width
            for col in ws.columns:
                max_length = 0
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)
            
            filename = f"laporan_peminjaman_{date.today()}.xlsx"
            wb.save(filename)
            msg.showinfo("Berhasil", f"File berhasil disimpan: {filename}")
            os.startfile(filename)
        except ImportError:
            msg.showerror("Error", "Library openpyxl belum terinstall.\nJalankan: pip install openpyxl")
        except Exception as e:
            msg.showerror("Error", f"Gagal export: {str(e)}")

    def _export_pengembalian_excel(self):
        if not hasattr(self, '_data_pengembalian'):
            msg.showwarning("Peringatan", "Tampilkan data terlebih dahulu")
            return
        
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Laporan Pengembalian"
            
            headers = ["ID", "Peminjam", "Email", "Tgl Pinjam", "Tgl Kembali Rencana", "Tgl Kembali Aktual", "Kondisi", "Denda", "Verifikator"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(1, col, header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            for row, d in enumerate(self._data_pengembalian, 2):
                ws.cell(row, 1, d['id'])
                ws.cell(row, 2, d['nama_peminjam'])
                ws.cell(row, 3, d.get('email', '-'))
                ws.cell(row, 4, str(d['tanggal_pinjam']))
                ws.cell(row, 5, str(d['tanggal_kembali']))
                ws.cell(row, 6, str(d['tanggal_kembali_aktual']))
                ws.cell(row, 7, d['kondisi_alat'])
                ws.cell(row, 8, d['total_denda'])
                ws.cell(row, 9, d.get('nama_verifikator', 'Belum'))
            
            for col in ws.columns:
                max_length = 0
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)
            
            filename = f"laporan_pengembalian_{date.today()}.xlsx"
            wb.save(filename)
            msg.showinfo("Berhasil", f"File berhasil disimpan: {filename}")
            os.startfile(filename)
        except ImportError:
            msg.showerror("Error", "Library openpyxl belum terinstall.\nJalankan: pip install openpyxl")
        except Exception as e:
            msg.showerror("Error", f"Gagal export: {str(e)}")

    def _export_denda_excel(self):
        if not hasattr(self, '_data_denda'):
            msg.showwarning("Peringatan", "Tampilkan data terlebih dahulu")
            return
        
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Laporan Denda"
            
            headers = ["ID", "Peminjam", "Email", "Jumlah", "Status", "Tgl Bayar", "Metode Bayar", "Keterangan"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(1, col, header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            for row, d in enumerate(self._data_denda, 2):
                ws.cell(row, 1, d['id'])
                ws.cell(row, 2, d['nama_peminjam'])
                ws.cell(row, 3, d.get('email', '-'))
                ws.cell(row, 4, d['jumlah'])
                ws.cell(row, 5, d['status'])
                ws.cell(row, 6, str(d.get('tanggal_bayar', '-')))
                ws.cell(row, 7, d.get('metode_bayar', '-'))
                ws.cell(row, 8, d.get('keterangan', '-'))
            
            for col in ws.columns:
                max_length = 0
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)
            
            filename = f"laporan_denda_{date.today()}.xlsx"
            wb.save(filename)
            msg.showinfo("Berhasil", f"File berhasil disimpan: {filename}")
            os.startfile(filename)
        except ImportError:
            msg.showerror("Error", "Library openpyxl belum terinstall.\nJalankan: pip install openpyxl")
        except Exception as e:
            msg.showerror("Error", f"Gagal export: {str(e)}")

    # Export PDF methods
    def _export_peminjaman_pdf(self):
        if not hasattr(self, '_data_peminjaman'):
            msg.showwarning("Peringatan", "Tampilkan data terlebih dahulu")
            return
        
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import cm
            from reportlab.lib.enums import TA_CENTER, TA_LEFT
            
            filename = f"laporan_peminjaman_{date.today()}.pdf"
            doc = SimpleDocTemplate(filename, pagesize=landscape(A4))
            elements = []
            
            # Styles
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                textColor=colors.HexColor('#1E40AF'),
                spaceAfter=30,
                alignment=TA_CENTER
            )
            
            # Title
            elements.append(Paragraph("LAPORAN PEMINJAMAN ALAT", title_style))
            elements.append(Paragraph(f"Periode: {self.pem_tgl_dari.get()} s/d {self.pem_tgl_sampai.get()}", 
                                     ParagraphStyle('subtitle', parent=styles['Normal'], alignment=TA_CENTER)))
            elements.append(Spacer(1, 0.5*cm))
            
            # Table data
            data = [['ID', 'Peminjam', 'Tgl Pinjam', 'Tgl Kembali', 'Status', 'Alat']]
            for d in self._data_peminjaman:
                data.append([
                    str(d['id']),
                    d['nama_peminjam'][:20],
                    str(d['tanggal_pinjam']),
                    str(d['tanggal_kembali']),
                    d['status'],
                    (d.get('alat_list', '-') or '-')[:30]
                ])
            
            # Create table
            table = Table(data, colWidths=[1.5*cm, 4*cm, 3*cm, 3*cm, 2.5*cm, 6*cm])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E40AF')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
            ]))
            
            elements.append(table)
            doc.build(elements)
            
            msg.showinfo("Berhasil", f"File PDF berhasil disimpan: {filename}")
            os.startfile(filename)
        except ImportError:
            msg.showerror("Error", "Library reportlab belum terinstall.\nJalankan: pip install reportlab")
        except Exception as e:
            msg.showerror("Error", f"Gagal export PDF: {str(e)}")

    def _export_pengembalian_pdf(self):
        if not hasattr(self, '_data_pengembalian'):
            msg.showwarning("Peringatan", "Tampilkan data terlebih dahulu")
            return
        
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import cm
            from reportlab.lib.enums import TA_CENTER
            
            filename = f"laporan_pengembalian_{date.today()}.pdf"
            doc = SimpleDocTemplate(filename, pagesize=landscape(A4))
            elements = []
            
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                textColor=colors.HexColor('#1E40AF'),
                spaceAfter=30,
                alignment=TA_CENTER
            )
            
            elements.append(Paragraph("LAPORAN PENGEMBALIAN ALAT", title_style))
            elements.append(Paragraph(f"Periode: {self.pen_tgl_dari.get()} s/d {self.pen_tgl_sampai.get()}", 
                                     ParagraphStyle('subtitle', parent=styles['Normal'], alignment=TA_CENTER)))
            elements.append(Spacer(1, 0.5*cm))
            
            data = [['ID', 'Peminjam', 'Tgl Kembali', 'Kondisi', 'Denda', 'Verifikator']]
            for d in self._data_pengembalian:
                data.append([
                    str(d['id']),
                    d['nama_peminjam'][:20],
                    str(d['tanggal_kembali_aktual']),
                    d['kondisi_alat'],
                    rupiah(d['total_denda']),
                    (d.get('nama_verifikator') or 'Belum')[:20]
                ])
            
            table = Table(data, colWidths=[1.5*cm, 4*cm, 3*cm, 2.5*cm, 3*cm, 4*cm])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E40AF')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
            ]))
            
            elements.append(table)
            doc.build(elements)
            
            msg.showinfo("Berhasil", f"File PDF berhasil disimpan: {filename}")
            os.startfile(filename)
        except ImportError:
            msg.showerror("Error", "Library reportlab belum terinstall.\nJalankan: pip install reportlab")
        except Exception as e:
            msg.showerror("Error", f"Gagal export PDF: {str(e)}")

    def _export_denda_pdf(self):
        if not hasattr(self, '_data_denda'):
            msg.showwarning("Peringatan", "Tampilkan data terlebih dahulu")
            return
        
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import cm
            from reportlab.lib.enums import TA_CENTER
            
            filename = f"laporan_denda_{date.today()}.pdf"
            doc = SimpleDocTemplate(filename, pagesize=landscape(A4))
            elements = []
            
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                textColor=colors.HexColor('#1E40AF'),
                spaceAfter=30,
                alignment=TA_CENTER
            )
            
            elements.append(Paragraph("LAPORAN DENDA", title_style))
            elements.append(Paragraph(f"Periode: {self.den_tgl_dari.get()} s/d {self.den_tgl_sampai.get()}", 
                                     ParagraphStyle('subtitle', parent=styles['Normal'], alignment=TA_CENTER)))
            elements.append(Spacer(1, 0.5*cm))
            
            data = [['ID', 'Peminjam', 'Jumlah', 'Status', 'Tgl Bayar', 'Keterangan']]
            for d in self._data_denda:
                data.append([
                    str(d['id']),
                    d['nama_peminjam'][:20],
                    rupiah(d['jumlah']),
                    d['status'],
                    str(d.get('tanggal_bayar', '-')),
                    (d.get('keterangan', '-') or '-')[:30]
                ])
            
            table = Table(data, colWidths=[1.5*cm, 4*cm, 3*cm, 2.5*cm, 3*cm, 6*cm])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E40AF')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
            ]))
            
            elements.append(table)
            doc.build(elements)
            
            msg.showinfo("Berhasil", f"File PDF berhasil disimpan: {filename}")
            os.startfile(filename)
        except ImportError:
            msg.showerror("Error", "Library reportlab belum terinstall.\nJalankan: pip install reportlab")
        except Exception as e:
            msg.showerror("Error", f"Gagal export PDF: {str(e)}")
